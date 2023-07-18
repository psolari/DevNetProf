import webbrowser
import random
from openpyxl import load_workbook


random_number = random.randint(1, 100)
if random_number <= 65:
    random_number = random.randint(18, 36)
else:
    random_number = random.randint(7, 17)


workbook = load_workbook('spreadsheet2.xlsx')
sheet = workbook.active
next_row = sheet.max_row + 1
sheet[f'A{next_row}'] = random_number
workbook.save('spreadsheet2.xlsx')
url = f"https://www.examtopics.com/exams/cisco/350-901/view/{random_number}"
webbrowser.open(url)